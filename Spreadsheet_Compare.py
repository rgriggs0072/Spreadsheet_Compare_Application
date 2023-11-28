import os
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from io import BytesIO

# Function to compare Excel sheets and create a new workbook
def compare_and_create_workbook(file1, file2):
    

    # Read the uploaded file into a pandas DataFrame
    df1 = pd.read_excel(file1)

    # Load the Excel file with openpyxl
    workbook1 = load_workbook(file1)

    # Get the names of all sheets in the workbook
    sheet_names1 = workbook1.sheetnames

    # Assuming there's only one sheet, you can access its name like this
    sheet_name1 = sheet_names1[0]

    # Read the uploaded file into a pandas DataFrame
    df2 = pd.read_excel(file2)

    # Load the Excel file with openpyxl
    workbook2 = load_workbook(file2)

    # Get the names of all sheets in the workbook
    sheet_names2 = workbook2.sheetnames

    # Assuming there's only one sheet, you can access its name like this
    sheet_name2 = sheet_names2[0]

    # Determine the column order based on the dataframes
    column_order = list(df1.columns)  # Use the columns from Sheet1 as the order

    # Create a new Excel workbook
    compared_wb = Workbook()

    # Initialize sheets
    sheet_names = [
        "Mismatch between Source 1 and Source 2",
        "Rows in Sheet1 not in Sheet2",
        "Rows in Sheet2 not in Sheet1",
    ]

    for sheet_name in sheet_names:
        compared_wb.create_sheet(sheet_name)

    # Find records with matching record IDs in both sheets
    matching_record_ids = set(df1["emp_id"]).intersection(df2["emp_id"])

    # Initialize lists to store rows for each sheet
    rows_for_mismatch = []
    rows_for_sheet1_not_in_sheet2 = []
    rows_for_sheet2_not_in_sheet1 = []

    # Compare data for matching record IDs
    for record_id in matching_record_ids:
        # Get rows with the same record ID from both dataframes
        row_in_df1 = df1[df1["emp_id"] == record_id]
        row_in_df2 = df2[df2["emp_id"] == record_id]

        # Check if the data is different
        differing_columns = {}

        if not row_in_df1.empty and not row_in_df2.empty:
         for col in column_order:
            value1 = row_in_df1.iloc[0][col]
            value2 = row_in_df2.iloc[0][col]

            # Convert to string if it's a float
            if isinstance(value1, float):
                value1 = str(value1)
            if isinstance(value2, float):
                value2 = str(value2)

            if value1 != value2:
                differing_columns[col] = {
                    "sheet 1": str(value1),
                    "sheet 2": str(value2),
                }

        if differing_columns:
            rows_for_mismatch.append({"emp_id": record_id, **differing_columns})

    # Get the worksheet for "Mismatch between Source 1 and Source 2"
    ws_mismatch = compared_wb["Mismatch between Source 1 and Source 2"]

    # Write the headers to the worksheet
    headers = ["emp_id", "Field", "Sheet 1", "Sheet 2"]
    ws_mismatch.append(headers)

    # Write the mismatched data to the worksheet
    for row in rows_for_mismatch:
        emp_id = row["emp_id"]
        for col, values in row.items():
            if col != "emp_id":
                field = col
                sheet1_value = values["sheet 1"]
                sheet2_value = values["sheet 2"]
                ws_mismatch.append([emp_id, field, sheet1_value, sheet2_value])

    # Find rows in Sheet1 not in Sheet2
    rows_in_sheet1_not_in_sheet2 = []

    for _, row in df1.iterrows():
        if row["emp_id"] not in matching_record_ids:
            rows_in_sheet1_not_in_sheet2.append(row.to_dict())

    # Find rows in Sheet2 not in Sheet1
    rows_in_sheet2_not_in_sheet1 = []

    for _, row in df2.iterrows():
        if row["emp_id"] not in matching_record_ids:
            rows_in_sheet2_not_in_sheet1.append(row.to_dict())

    # Create DataFrames for rows not in Sheet2 and Sheet1
    df_rows_in_sheet1_not_in_sheet2 = pd.DataFrame(rows_in_sheet1_not_in_sheet2)
    df_rows_in_sheet2_not_in_sheet1 = pd.DataFrame(rows_in_sheet2_not_in_sheet1)

    # Get the worksheets for other sheets
    ws_sheet1_not_in_sheet2 = compared_wb["Rows in Sheet1 not in Sheet2"]
    ws_sheet2_not_in_sheet1 = compared_wb["Rows in Sheet2 not in Sheet1"]

    # Write DataFrames to corresponding worksheets
    for idx, df in enumerate([df_rows_in_sheet1_not_in_sheet2, df_rows_in_sheet2_not_in_sheet1]):
        ws = [ws_sheet1_not_in_sheet2, ws_sheet2_not_in_sheet1][idx]
        for r_idx, row in df.iterrows():
            values = [row[col] for col in column_order]
            if r_idx == 0:
                ws.append(column_order)  # Write the column names in the correct order
            ws.append(values)
            
    # Save the Excel workbook to a BytesIO buffer
    buffer = BytesIO()
    compared_wb.save(buffer)
    
    compared_wb.close()

    return buffer
  

 # Streamlit app
st.title("Excel Sheet Comparison Application")

# Upload the first spreadsheet
uploaded_file1 = st.file_uploader("Upload Excel File 1 (Sheet1)", type=["xlsx"])

# Upload the second spreadsheet
uploaded_file2 = st.file_uploader("Upload Excel File 2 (Sheet1)", type=["xlsx"])

# Compare Excel sheets and create a workbook with separate sheets
if st.button("Compare and Create Workbook"):
    if uploaded_file1 and uploaded_file2:
        compared_buffer = compare_and_create_workbook(uploaded_file1, uploaded_file2)
        st.success("Comparison completed. Download the compared Excel workbook below:")

        # Specify the file type as 'xlsx' in the download button
        st.download_button(
            label="Download Compared Excel Workbook",
            data=compared_buffer.getvalue(),
            key='compared_excel',
            file_name="compared_results.xlsx"
        )
    else:
        st.warning("Please upload both Excel files for comparison.")